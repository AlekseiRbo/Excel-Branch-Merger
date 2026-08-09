from pathlib import Path
import openpyxl
import pandas as pd
from src.excel_branch_merger.merger import process_folder

CONFIG = {
    "canonical_columns": {"customer_name": ["Customer"], "sale_date": ["Date"], "amount": ["Amount"], "invoice_number": ["Invoice"]},
    "required_fields": ["customer_name", "sale_date", "amount"],
    "duplicate_key": ["customer_name", "sale_date", "amount", "invoice_number"],
    "date_formats": ["%Y-%m-%d"],
}

def test_all_invalid_rows_still_export_both_workbooks(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"; input_dir.mkdir(); output_dir = tmp_path / "output"
    pd.DataFrame({"Customer": ["Alice"], "Date": ["not-a-date"], "Amount": ["bad"], "Invoice": ["A-1"]}).to_excel(input_dir / "bad.xlsx", index=False)
    result = process_folder(input_dir, output_dir, CONFIG)
    assert result.valid_rows == 0
    assert result.error_rows >= 1
    for path in (result.report_path, result.error_path):
        workbook = openpyxl.load_workbook(path, read_only=True)
        workbook.close()
    consolidated = pd.read_excel(result.report_path, sheet_name="Consolidated")
    errors = pd.read_excel(result.error_path, sheet_name="Errors")
    assert consolidated.empty
    assert len(errors) == 1


def test_all_valid_rows_allow_empty_error_report(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"; input_dir.mkdir(); output_dir = tmp_path / "output"
    pd.DataFrame({
        "Customer": ["Alice"], "Date": ["2026-08-01"],
        "Amount": [100], "Invoice": ["A-1"],
    }).to_excel(input_dir / "good.xlsx", index=False)
    result = process_folder(input_dir, output_dir, CONFIG)
    errors = pd.read_excel(result.error_path, sheet_name="Errors")
    assert errors.empty
    workbook = openpyxl.load_workbook(result.error_path, read_only=True)
    assert "Errors" in workbook.sheetnames
    workbook.close()


def test_header_only_workbook_exports_without_crashing(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"; input_dir.mkdir(); output_dir = tmp_path / "output"
    pd.DataFrame(columns=["Customer", "Date", "Amount", "Invoice"]).to_excel(
        input_dir / "headers.xlsx", index=False
    )
    result = process_folder(input_dir, output_dir, CONFIG)
    workbook = openpyxl.load_workbook(result.report_path, read_only=True)
    assert "Consolidated" in workbook.sheetnames and "Summary" in workbook.sheetnames
    workbook.close()
