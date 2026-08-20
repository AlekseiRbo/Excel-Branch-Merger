from pathlib import Path

import openpyxl
import pandas as pd

from src.excel_branch_merger.merger import process_folder


def _config(*, include_source_lineage: bool = True) -> dict:
    return {
        "input": {
            "csv": {
                "encoding": "utf-8-sig",
                "delimiter": ",",
            },
            "unmapped_columns": "preserve",
        },
        "fields": {
            "customer_name": {
                "aliases": ["Customer"],
                "required": True,
                "type": "text",
                "case": "preserve",
            },
            "email": {
                "aliases": ["Email"],
                "required": False,
                "type": "email",
                "case": "lower",
            },
            "amount": {
                "aliases": ["Amount"],
                "required": True,
                "type": "number",
            },
            "invoice_date": {
                "aliases": ["Date"],
                "required": True,
                "type": "date",
                "formats": ["%Y-%m-%d"],
            },
        },
        "deduplication": {
            "keys": ["email", "invoice_date"],
            "keep": "first",
        },
        "missing_values": ["", "N/A", "NA", "null"],
        "output": {
            "include_source_lineage": include_source_lineage,
        },
    }


def _write_input(path: Path) -> None:
    pd.DataFrame(
        [
            {
                "Customer": "Alpha",
                "Email": "alpha@example.com",
                "Amount": "100",
                "Date": "2026-08-20",
            },
            {
                "Customer": "Broken",
                "Email": "broken@example.com",
                "Amount": "not-a-number",
                "Date": "2026-08-20",
            },
        ]
    ).to_csv(
        path,
        index=False,
        encoding="utf-8-sig",
    )


def test_v14_uses_commercial_output_names_and_sheets(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _write_input(input_dir / "sales.csv")

    result = process_folder(
        input_dir,
        output_dir,
        _config(),
    )

    assert result.report_path.name == "Consolidated.xlsx"
    assert result.error_path.name == "Errors.xlsx"
    assert result.log_path.name == "processing.log"

    assert result.report_path.exists()
    assert result.error_path.exists()
    assert result.log_path.exists()

    report_book = openpyxl.load_workbook(
        result.report_path,
        read_only=True,
    )
    error_book = openpyxl.load_workbook(
        result.error_path,
        read_only=True,
    )

    try:
        assert report_book.sheetnames == ["Data", "Summary"]
        assert error_book.sheetnames == ["Errors"]
    finally:
        report_book.close()
        error_book.close()


def test_v14_summary_matches_processing_result(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _write_input(input_dir / "sales.csv")

    result = process_folder(
        input_dir,
        output_dir,
        _config(),
    )

    summary = pd.read_excel(
        result.report_path,
        sheet_name="Summary",
    )
    values = dict(zip(summary["Metric"], summary["Value"]))

    assert values["Files discovered"] == result.files_discovered
    assert values["Files succeeded"] == result.files_succeeded
    assert values["Files failed"] == result.files_failed
    assert values["Total input rows"] == result.total_input_rows
    assert values["Valid rows"] == result.valid_rows
    assert values["Invalid rows"] == result.invalid_rows
    assert values["Duplicate rows"] == result.duplicate_rows
    assert values["Incomplete dedup key rows"] == result.incomplete_dedup_key_rows
    assert values["Total rejected rows"] == result.total_rejected_rows

    assert result.total_input_rows == (
        result.valid_rows + result.invalid_rows + result.duplicate_rows
    )


def test_v14_lineage_is_included_when_enabled(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _write_input(input_dir / "sales.csv")

    result = process_folder(
        input_dir,
        output_dir,
        _config(include_source_lineage=True),
    )

    data = pd.read_excel(
        result.report_path,
        sheet_name="Data",
    )
    errors = pd.read_excel(
        result.error_path,
        sheet_name="Errors",
    )

    lineage = {
        "source_file",
        "source_sheet",
        "source_row",
    }

    assert lineage <= set(data.columns)
    assert lineage <= set(errors.columns)

    assert data.loc[0, "source_file"] == "sales.csv"
    assert data.loc[0, "source_sheet"] == "CSV"
    assert data.loc[0, "source_row"] == 2


def test_v14_lineage_is_omitted_when_disabled(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _write_input(input_dir / "sales.csv")

    result = process_folder(
        input_dir,
        output_dir,
        _config(include_source_lineage=False),
    )

    data = pd.read_excel(
        result.report_path,
        sheet_name="Data",
    )
    errors = pd.read_excel(
        result.error_path,
        sheet_name="Errors",
    )

    lineage = {
        "source_file",
        "source_sheet",
        "source_row",
    }

    assert lineage.isdisjoint(data.columns)
    assert lineage.isdisjoint(errors.columns)


def test_v14_output_files_are_not_rediscovered_as_inputs(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()

    _write_input(input_dir / "sales.csv")

    pd.DataFrame({"ignore": [1]}).to_excel(
        input_dir / "Consolidated.xlsx",
        index=False,
    )
    pd.DataFrame({"ignore": [1]}).to_excel(
        input_dir / "Errors.xlsx",
        index=False,
    )

    output_dir = tmp_path / "output"

    result = process_folder(
        input_dir,
        output_dir,
        _config(),
    )

    assert result.files_discovered == 1
    assert result.files_succeeded == 1
